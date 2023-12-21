import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

# Função para verificar se a planilha já existe
def planilha_existe(nome_arquivo):
    return os.path.exists(nome_arquivo)

tecnico = input("Informe o nome do Técnico: ")
numero_os = int(input("Informe o número de ordens de serviços finalizadas por período: "))
valorTotal_os = float(input("Informe o valor em reais das ordens de serviços finalizadas: "))

mediaTicket = valorTotal_os / numero_os
mediaDia = numero_os / 30
mediaDia2 = valorTotal_os / 30

print()
print(f"O Técnico {tecnico} tem um ticket médio por O.S de {mediaTicket:.2f} reais.")
print(f"Ele possui uma média de {mediaDia:.2f} O.S's liberadas por dia.")
print(f"O valor médio de faturamento de O.S's por dia é de {mediaDia2:.2f} reais.")
print()
print()

input("Pressione Enter para fechar a janela.")

nome_arquivo = f"resultados_{tecnico}.xlsx"

# Se a planilha existir, carregá-la
if planilha_existe(nome_arquivo):
    # Carregar a planilha existente
    wb = openpyxl.load_workbook(nome_arquivo)
    sheet = wb.active
else:
    # Se a planilha não existir, criar uma nova
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Resultados"

    # Definir os cabeçalhos da tabela e aplicar formatação
    headers = ["Data", "Técnico", "Ticket Médio por O.S (R$)", "Média de O.S's por Dia", "Valor Médio de Faturamento por Dia (R$)"]
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="b2ebf2", end_color="b2ebf2", fill_type="solid")

# Obter a data atual
data_atual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# Preencher os resultados nas células
data = [data_atual, tecnico, round(mediaTicket, 2), round(mediaDia, 2), round(mediaDia2, 2)]
sheet.append(data)

# Aplicar formatação às células de dados
for row in sheet.iter_rows(min_row=sheet.max_row, max_row=sheet.max_row):
    for cell in row:
        cell.alignment = Alignment(horizontal="center")

# Destacar as células de média de O.S's por Dia e Valor Médio de Faturamento por Dia com cores diferentes
media_os_celula = sheet.cell(row=sheet.max_row, column=4)
media_faturamento_celula = sheet.cell(row=sheet.max_row, column=5)

media_os_celula.fill = PatternFill(start_color="c8e6c9", end_color="c8e6c9", fill_type="solid")
media_faturamento_celula.fill = PatternFill(start_color="ffccbc", end_color="ffccbc", fill_type="solid")

# Redimensionar as colunas para ajustar o conteúdo
for column_cells in sheet.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

# Salvar a planilha em um arquivo
wb.save(nome_arquivo)

# Fechar a planilha
wb.close()

print(f"Os resultados foram armazenados em {nome_arquivo}.")






