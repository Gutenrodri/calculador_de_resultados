import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os

# Função para verificar se a planilha já existe
def planilha_existe(nome_arquivo):
    return os.path.exists(nome_arquivo)

def calcular_media():
    tecnico = entry_tecnico.get()
    numero_os = int(entry_numero_os.get())
    valorTotal_os = float(entry_valorTotal_os.get())

    mediaTicket = valorTotal_os / numero_os
    mediaDia = numero_os / 30
    mediaDia2 = valorTotal_os / 30

    result_label.config(text=f"O Técnico {tecnico} tem um ticket médio por O.S de {mediaTicket:.2f} reais.\n"
                             f"Ele possui uma média de {mediaDia:.2f} O.S's liberadas por dia.\n"
                             f"O valor médio de faturamento de O.S's por dia é de {mediaDia2:.2f} reais.")

    salvar_resultados(tecnico, mediaTicket, mediaDia, mediaDia2)

def salvar_resultados(tecnico, mediaTicket, mediaDia, mediaDia2):
    nome_arquivo = f"resultados_{tecnico}.xlsx"

    # Se a planilha existir, carregá-la
    if planilha_existe(nome_arquivo):
        # Carregar a planilha existente
        wb = openpyxl.load_workbook(nome_arquivo)
        sheet = wb.active
    else:
        # Se a planilha não existir, criar uma nova
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Resultados"

        # Definir os cabeçalhos da tabela e aplicar formatação
        headers = ["Data", "Técnico", "Ticket Médio por O.S (R$)", "Média de O.S's por Dia", "Valor Médio de Faturamento por Dia (R$)"]
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color="b2ebf2", end_color="b2ebf2", fill_type="solid")

    # Obter a data atual
    data_atual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Preencher os resultados nas células
    data = [data_atual, tecnico, round(mediaTicket, 2), round(mediaDia, 2), round(mediaDia2, 2)]
    sheet.append(data)

    # Aplicar formatação às células de dados
    for row in sheet.iter_rows(min_row=sheet.max_row, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    # Destacar as células de média de O.S's por Dia e Valor Médio de Faturamento por Dia com cores diferentes
    media_os_celula = sheet.cell(row=sheet.max_row, column=4)
    media_faturamento_celula = sheet.cell(row=sheet.max_row, column=5)

    media_os_celula.fill = PatternFill(start_color="c8e6c9", end_color="c8e6c9", fill_type="solid")
    media_faturamento_celula.fill = PatternFill(start_color="ffccbc", end_color="ffccbc", fill_type="solid")

    # Redimensionar as colunas para ajustar o conteúdo
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Salvar a planilha em um arquivo
    wb.save(nome_arquivo)

    # Fechar a planilha
    wb.close()

    messagebox.showinfo("Salvo", f"Os resultados foram armazenados em {nome_arquivo}.")

# Criar janela principal
root = tk.Tk()
root.title("Calculadora de Média")
root.geometry("400x300")

# Cores para a interface
cor_fundo = "#F0F0F0"
cor_botao = "#4CAF50"
cor_label = "#333333"

# Widgets
root.configure(bg=cor_fundo)

label_tecnico = tk.Label(root, text="Informe o nome do Técnico:", bg=cor_fundo, fg=cor_label)
label_tecnico.pack()

entry_tecnico = tk.Entry(root)
entry_tecnico.pack()

label_numero_os = tk.Label(root, text="Informe o número de ordens de serviços finalizadas por período:", bg=cor_fundo, fg=cor_label)
label_numero_os.pack()

entry_numero_os = tk.Entry(root)
entry_numero_os.pack()

label_valorTotal_os = tk.Label(root, text="Informe o valor em reais das ordens de serviços finalizadas:", bg=cor_fundo, fg=cor_label)
label_valorTotal_os.pack()

entry_valorTotal_os = tk.Entry(root)
entry_valorTotal_os.pack()

calculate_button = tk.Button(root, text="Calcular Média e Salvar", command=calcular_media, bg=cor_botao, fg="white")
calculate_button.pack()

result_label = tk.Label(root, text="", bg=cor_fundo, fg=cor_label)
result_label.pack()

root.mainloop()
